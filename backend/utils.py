import os
import logging
from google.cloud import storage
import json
import vertexai
from vertexai.preview import rag
from vertexai.preview.generative_models import GenerativeModel, Tool
import tempfile
from concurrent.futures import ThreadPoolExecutor
from dotenv import load_dotenv

import io
import PyPDF2
import docx2txt
import openpyxl
from typing import Dict, Any

load_dotenv()

GCS_BUCKET_NAME = os.environ.get("GCS_BUCKET_NAME", "NO BUCKET NAME")
PROJECT_ID = os.environ.get("PROJECT_ID", "your-project-id")
LOCATION = os.environ.get("LOCATION", "us-central1")
# Initialize Vertex AI API once per session
vertexai.init(project=PROJECT_ID, location=LOCATION)

# Global dictionary to store corpus name and its identifier. It can also be a database if needed
corpus_registry = {}

def setup_logging():
    log_level = os.environ.get("LOG_LEVEL", "INFO").upper()
    numeric_level = getattr(logging, log_level, None)
    if not isinstance(numeric_level, int):
        raise ValueError(f"Invalid log level: {log_level}")
    logging.basicConfig(level=numeric_level, format='%(asctime)s - %(levelname)s - %(message)s')


def upload_to_gcs(bucket_name, filename, content, content_type="application/octet-stream"):
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(filename)
    blob.upload_from_string(content, content_type=content_type)


def download_from_gcs(bucket_name, filename):
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(filename)
    if blob.exists():
        content = blob.download_as_string()
        logging.info(f"Downloaded {filename} from GCS bucket {bucket_name}")
        return content
    else:
        logging.info(f"{filename} not found in GCS bucket {bucket_name}")
        return None


def save_scraped_data_to_gcs(scraped_data, bucket_name, filename):
    try:
        content = json.dumps(scraped_data)
        upload_to_gcs(bucket_name, filename, content, content_type="application/json")
        logging.info("Successfully saved scraped data to GCS")
    except Exception as e:
        logging.error(f"Error saving scraped data to GCS: {e}")


def load_scraped_data_from_gcs(bucket_name, filename):
    try:
        content = download_from_gcs(bucket_name, filename)
        if content:
            return json.loads(content)
        else:
            return None
    except Exception as e:
        logging.error(f"Error loading scraped data from GCS: {e}")
        return None


def create_rag_corpus(display_name, description):
    embedding_model = "text-embedding-004"

    try:
        corpora = rag.list_corpora()
        logging.info(f"Existing corpora: {corpora}")
        for corpus in corpora:
            if corpus.display_name == display_name:
                logging.info(f"RAG Corpus {display_name} already exists: {corpus.name}")
                return corpus.name  # Return the existing corpus name

        logging.info(f"Creating RAG Corpus: {display_name}")
        corpus = rag.create_corpus(
            display_name=display_name,
            description=description
        )
        logging.info(f"RAG Corpus created: {corpus.name}")
        return corpus.name
    except Exception as e:
        logging.error(f"Error creating/checking RAG corpus: {e}")
        return False


def import_files_to_corpus(corpus_name, paths, chunk_size=512, chunk_overlap=100, max_embedding_requests_per_min=900):
    try:
        response = rag.import_files(
            corpus_name=corpus_name,
            paths=paths,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            max_embedding_requests_per_min=max_embedding_requests_per_min,
        )
        logging.info(f"Imported {response.imported_rag_files_count} files to {corpus_name}.")
    except Exception as e:
        logging.error(f"Error uploading documents to RAG corpus: {e}")


def cleanup_gcs_files(bucket_name, gcs_paths):
    try:
        client = storage.Client()
        bucket = client.bucket(bucket_name)
        for gcs_path in gcs_paths:
            blob_name = gcs_path.replace(f"gs://{bucket_name}/", "")
            blob = bucket.blob(blob_name)
            if blob.exists():
                blob.delete()
                logging.info(f"Deleted GCS file: {gcs_path}")
            else:
                logging.warning(f"GCS file not found: {gcs_path}")
    except Exception as e:
        logging.error(f"Error during GCS cleanup: {e}")


def cleanup_gcs_bucket_parallel(bucket_name: str, max_workers: int = 10) -> None:
    try:
        client = storage.Client()
        bucket = client.bucket(bucket_name)
        blobs = list(bucket.list_blobs())

        if not blobs:
            logging.info(f"No files found in bucket '{bucket_name}'.")
            return

        def delete_blob(blob):
            try:
                blob.delete()
                logging.info(f"Deleted GCS file: {blob.name}")
            except Exception as e:
                logging.error(f"Error deleting blob {blob.name}: {e}")

        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            executor.map(delete_blob, blobs)

        logging.info(f"All files in bucket '{bucket_name}' have been deleted.")
    except Exception as e:
        logging.error(f"Error during GCS bucket cleanup: {e}")


def get_relevant_corpora(query):
    possible_keys = list(corpus_registry.keys())
    if not possible_keys:
        return []

    model = GenerativeModel(model_name="gemini-2.0-flash-exp")

    prompt = f"""
Given this user query: "{query}"
You have a list of corpora: {', '.join(possible_keys)}.
Which of these corpora are relevant?
Return all that apply, as a comma-separated list with no additional text.
If none apply, return "none".
"""
    response = model.generate_content(prompt)
    classification = response.text.strip().lower()
    logging.info(f"Multi-corpus classification result: '{classification}'")

    if classification == "none":
        return []

    chosen_keys = [k.strip() for k in classification.split(",")]
    chosen_keys = [ck for ck in chosen_keys if ck in corpus_registry]
    relevant_corpus_list = [corpus_registry[ck] for ck in chosen_keys]
    return relevant_corpus_list


def generate_rag_response(query: str, mode: str = "auto", manual_corpora=None):
    """
    Generate a response from the RAG system. If mode="auto", it will
    detect relevant corpora automatically. If mode="manual", it will
    ONLY search within the user-provided corpora (list of display_names).
    """
    # If manual mode, user has provided a list of display_names
    # which we look up in corpus_registry to get the full resource names
    if mode == "manual" and manual_corpora:
        corpora_list = []
        for display_name in manual_corpora:
            if display_name in corpus_registry:
                corpora_list.append(corpus_registry[display_name])
        if not corpora_list:
            return {
                "status": "Error",
                "response": "No valid corpora selected or they are not registered.",
            }
    else:
        # "auto" mode
        corpora_list = get_relevant_corpora(query)

    if not corpora_list:
        return {
            "status": "Error",
            "response": "No relevant documentation found.",
        }

    all_retrieved_docs = []
    for corpus_name in corpora_list:
        try:
            retrieval_source = rag.retrieval_query(
                rag_resources=[rag.RagResource(
                    rag_corpus=corpus_name,
                )],
                text=query,
                similarity_top_k=5,
            )
            retrieved_context = " ".join(
                [context.text for context in retrieval_source.contexts.contexts]
            ).replace("\n", "")
            all_retrieved_docs.extend(retrieved_context)
        except Exception as e:
            logging.error(f"Error retrieving from corpus {corpus_name}: {e}")

    if not all_retrieved_docs:
        return {
            "status": "Error",
            "response": "No matching documents across all corpora.",
        }

    context_text = "\n\n".join(all_retrieved_docs)
    final_prompt = f"""
####CONTEXT START:
{context_text}

####CONTEXT END

Answer the following user query, use the provided context:

####USER QUERY:
{query}
"""

    rag_model = GenerativeModel(model_name="gemini-2.0-flash-exp")

    try:
        response = rag_model.generate_content(final_prompt)
        return {
            "status": "OK",
            "response": response.text,
            "corpus_used": corpora_list
        }
    except Exception as e:
        logging.error(f"Error in multi-corpus generation: {e}")
        return {
            "status": "Error",
            "response": "I encountered an error. Please try again later."
        }


def handle_new_documentation(url, display_name, description, scraped_data):
    corpus_name = create_rag_corpus(display_name=display_name, description=description)
    if not corpus_name:
        return {"status": "Error", "message": "Could not create the corpus"}

    paths = []
    gcs_paths = []
    try:
        for url, text in scraped_data.items():
            if not isinstance(text, str):
                logging.warning(f"Non-string content found for URL {url}. Skipping...")
                continue
            with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False, encoding="utf-8") as tmp_file:
                tmp_file.write(text)
                tmp_file_path = tmp_file.name
                paths.append(tmp_file_path)

            gcs_path = f"gs://{GCS_BUCKET_NAME}/{os.path.basename(tmp_file_path)}"
            upload_to_gcs(GCS_BUCKET_NAME, os.path.basename(tmp_file_path), text, content_type="text/plain")
            gcs_paths.append(gcs_path)

        if not gcs_paths:
            return {"status": "Error", "message": "No valid documentation to import"}

        batch_size = 25
        for i in range(0, len(gcs_paths), batch_size):
            batch = gcs_paths[i:i + batch_size]
            logging.info(f"Importing batch {i // batch_size + 1} with {len(batch)} files")
            import_files_to_corpus(corpus_name=corpus_name, paths=batch)
    finally:
        for path in paths:
            os.remove(path)
        cleanup_gcs_bucket_parallel(GCS_BUCKET_NAME)

    corpus_registry[display_name] = corpus_name
    return {"status": "OK", "message": "Documentation indexed successfully!", "corpus_name": corpus_name}


def load_corpus_registry():
    try:
        with open("corpus_registry.json", "r") as f:
            global corpus_registry
            corpus_registry = json.load(f)
    except FileNotFoundError:
        logging.info("Corpus registry file not found, starting with an empty registry.")
        corpus_registry = {}
    logging.info(f"Corpus registry loaded: {corpus_registry}")


def save_corpus_registry():
    with open("corpus_registry.json", "w") as f:
        json.dump(corpus_registry, f)
    logging.info(f"Corpus registry saved: {corpus_registry}")


def delete_corpora():
    try:
        corpora = rag.list_corpora()
        for corpus in corpora:
            rag.delete_corpus(corpus.name)
            logging.info(f"Deleted RAG corpus: {corpus.name}")
    except Exception as e:
        logging.error(f"Error deleting RAG corpora: {e}")


def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    ext = os.path.splitext(filename.lower())[1]

    if ext == ".pdf":
        try:
            with io.BytesIO(file_bytes) as pdf_stream:
                pdf_reader = PyPDF2.PdfReader(pdf_stream)
                all_text = []
                for page in pdf_reader.pages:
                    text = page.extract_text()
                    if text:
                        all_text.append(text)
            return "\n".join(all_text)
        except Exception as e:
            logging.error(f"Error parsing PDF {filename}: {e}")
            return ""

    elif ext == ".docx":
        try:
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
                tmp.write(file_bytes)
                tmp.flush()
                text = docx2txt.process(tmp.name)
            return text
        except Exception as e:
            logging.error(f"Error parsing DOCX {filename}: {e}")
            return ""

    elif ext == ".doc":
        logging.warning("DOC parsing not natively supported.")
        return ""

    elif ext == ".xlsx":
        try:
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(file_bytes)
                tmp.flush()
                wb = openpyxl.load_workbook(tmp.name, data_only=True)
                all_text = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        row_text = [str(cell) for cell in row if cell is not None]
                        if row_text:
                            all_text.append(" | ".join(row_text))
                return "\n".join(all_text)
        except Exception as e:
            logging.error(f"Error parsing XLSX {filename}: {e}")
            return ""

    elif ext == ".xls":
        logging.warning("XLS parsing not implemented.")
        return ""

    elif ext == ".txt":
        try:
            return file_bytes.decode("utf-8", errors="replace")
        except Exception as e:
            logging.error(f"Error parsing TXT {filename}: {e}")
            return ""
    else:
        logging.warning(f"Unsupported file format for {filename}")
        return ""
